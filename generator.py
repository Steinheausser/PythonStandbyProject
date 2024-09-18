import random
import logging
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Set up logging
log_filename = "shift_scheduler.log"
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
)


# Redirect stdout and stderr to logging
class LoggingStreamHandler:
    def __init__(self, level):
        self.level = level

    def write(self, message):
        if message.strip() != "":
            self.level(message.strip())

    def flush(self):
        pass

sys.stdout = LoggingStreamHandler(logging.info)
sys.stderr = LoggingStreamHandler(logging.error)

def rotate_list(lst, n):
    """Rotate a list by n positions."""
    return lst[n:] + lst[:n]

class ShiftScheduler:
    def __init__(self, start_date, end_date, names, holidays):
        self.start_date = start_date
        self.end_date = end_date
        self.names = names
        self.holidays = holidays
        self.schedule = {}
        self.assignments = {name: {'total': 0, 'special_days': 0, 'dates': [], 'last_assigned': None, 'weeks': {}} for name in names}
        self.special_days = self.calculate_special_days()
        self.total_shifts = ((self.end_date - self.start_date).days + 1) * 2

    def calculate_special_days(self):
        special_days = []
        current_date = self.start_date
        while current_date <= self.end_date:
            if current_date.weekday() >= 5 or current_date in self.holidays:
                special_days.append(current_date)
            current_date += timedelta(days=1)
        return special_days

    def is_special_day(self, date):
        return date.weekday() >= 5 or date in self.holidays

    def get_week_number(self, date):
        return (date - self.start_date).days // 7

    def is_available(self, name, date):
        week_number = self.get_week_number(date)
        return self.assignments[name]['weeks'].get(week_number, 0) < 2

    def is_consecutive(self, name, date):
        prev_day = date - timedelta(days=1)
        next_day = date + timedelta(days=1)
        return (prev_day in self.schedule and name in self.schedule[prev_day]) or \
               (next_day in self.schedule and name in self.schedule[next_day])

    def generate_schedule(self):
        logging.info("Generating initial schedule...")
        dates = [self.start_date + timedelta(days=i) for i in range((self.end_date - self.start_date).days + 1)]
        
        for current_date in dates:
            available_names = [name for name in self.names if not self.is_consecutive(name, current_date) and self.is_available(name, current_date)]
            if len(available_names) < 2:
                logging.warning(f"Not enough available names for date {current_date}. Attempting to relax constraints.")
                available_names = [name for name in self.names if not self.is_consecutive(name, current_date)]

            random.shuffle(available_names)

            shift = []
            for _ in range(2):  # Assign 2 people per day
                if not available_names:
                    logging.warning(f"No available names for date {current_date}. Choosing from all names.")
                    available_names = [name for name in self.names if name not in shift]
                    random.shuffle(available_names)

                name = available_names.pop(0)
                shift.append(name)
                week_number = self.get_week_number(current_date)
                self.assignments[name]['weeks'][week_number] = self.assignments[name]['weeks'].get(week_number, 0) + 1

            self.schedule[current_date] = shift

        logging.info("Initial schedule generated. Starting balancing process...")
        self.update_assignments()
        self.equalize_shifts()
        logging.info("Schedule generation and balancing completed.")

    def equalize_shifts(self, max_iterations=2000):
        logging.info("Starting to equalize shifts...")
        base_shifts = self.total_shifts // len(self.names)
        extra_shifts = self.total_shifts % len(self.names)

        target_shifts = {name: base_shifts + (1 if i < extra_shifts else 0) for i, name in enumerate(self.names)}
        target_special_days = len(self.special_days) * 2 // len(self.names)

        logging.info(f"Target shifts: {target_shifts}")
        logging.info(f"Target special days per person: {target_special_days}")

        iterations = 0
        while iterations < max_iterations:
            max_total = max(person['total'] for person in self.assignments.values())
            min_total = min(person['total'] for person in self.assignments.values())
            max_special = max(person['special_days'] for person in self.assignments.values())
            min_special = min(person['special_days'] for person in self.assignments.values())

            if max_total - min_total <= 1 and max_special - min_special <= 1:
                break

            for date in sorted(self.schedule.keys()):
                current_assignees = self.schedule[date]
                is_special = self.is_special_day(date)

                for i, name in enumerate(current_assignees):
                    if (self.assignments[name]['total'] > target_shifts[name] or 
                        (is_special and self.assignments[name]['special_days'] > target_special_days)):
                        
                        candidates = [n for n in self.names 
                                      if n not in current_assignees 
                                      and not self.is_consecutive(n, date)
                                      and self.is_available(n, date)
                                      and (self.assignments[n]['total'] < target_shifts[n] or
                                           (is_special and self.assignments[n]['special_days'] < target_special_days))]
                        
                        if candidates:
                            replacement = random.choice(candidates)
                            self.schedule[date][i] = replacement
                            self.assignments[name]['total'] -= 1
                            self.assignments[name]['dates'].remove(date)
                            self.assignments[replacement]['total'] += 1
                            self.assignments[replacement]['dates'].append(date)
                            if is_special:
                                self.assignments[name]['special_days'] -= 1
                                self.assignments[replacement]['special_days'] += 1
                            
                            # Update weeks
                            week_number = self.get_week_number(date)
                            self.assignments[name]['weeks'][week_number] -= 1
                            self.assignments[replacement]['weeks'][week_number] = self.assignments[replacement]['weeks'].get(week_number, 0) + 1

            iterations += 1

        logging.info(f"Finished equalizing shifts after {iterations} iterations.")
        self.update_assignments()

    def update_assignments(self):
        for name in self.names:
            self.assignments[name]['total'] = 0
            self.assignments[name]['special_days'] = 0
            self.assignments[name]['dates'] = []
            self.assignments[name]['last_assigned'] = None
            self.assignments[name]['weeks'] = {}

        for date, assignees in self.schedule.items():
            for name in assignees:
                self.assignments[name]['total'] += 1
                self.assignments[name]['dates'].append(date)
                if self.is_special_day(date):
                    self.assignments[name]['special_days'] += 1
                if self.assignments[name]['last_assigned'] is None or date > self.assignments[name]['last_assigned']:
                    self.assignments[name]['last_assigned'] = date
                week_number = self.get_week_number(date)
                self.assignments[name]['weeks'][week_number] = self.assignments[name]['weeks'].get(week_number, 0) + 1

    def print_schedule(self):
        print("Schedule by Date:")
        for date, shift in sorted(self.schedule.items()):
            print(f"{date.strftime('%Y-%m-%d')}: {', '.join(shift)}")

    def print_personal_schedules(self):
        print("\nSchedule by Person:")
        for name, stats in self.assignments.items():
            dates = ', '.join(date.strftime('%Y-%m-%d') for date in sorted(stats['dates']))
            print(f"{name}: {dates}")

    def print_statistics(self):
        print("\nAssignment Statistics:")
        for name, stats in self.assignments.items():
            print(f"{name}: Total: {stats['total']}, Special Days: {stats['special_days']}, Weeks: {len(stats['weeks'])}")

        print("\nOverall Statistics:")
        print(f"Total number of standbys: {self.total_shifts}")
        print(f"Total number of Special Days: {len(self.special_days)}")

    def export_to_excel(self, filename):
        wb = Workbook()

        # Shift Schedule sheet
        ws = wb.active
        ws.title = "Shift Schedule"
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        ws['A1'] = "Date"
        ws['B1'] = "Day"
        ws['C1'] = "Person 1"
        ws['D1'] = "Person 2"
        ws['E1'] = "Special Day"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Data for shift schedule
        for row, (date, shift) in enumerate(sorted(self.schedule.items()), start=2):
            ws.cell(row=row, column=1, value=date.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=date.strftime("%A"))  # Add day of the week
            for col, name in enumerate(shift, start=3):
                cell = ws.cell(row=row, column=col, value=name)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value="Yes" if self.is_special_day(date) else "No")
            ws.cell(row=row, column=5).border = border
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

        # Adjust column widths for shift schedule
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Personal Schedules sheet
        ws2 = wb.create_sheet("Personal Schedules")

        # Headers
        ws2['A1'] = "Person"
        ws2['B1'] = "Dates"
        ws2['C1'] = "Total Shifts"
        ws2['D1'] = "Special Days"
        ws2['E1'] = "Weeks"
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Data for personal schedules
        for row, (name, stats) in enumerate(self.assignments.items(), start=2):
            ws2.cell(row=row, column=1, value=name)
            ws2.cell(row=row, column=2, value=", ".join(date.strftime("%d/%m/%Y") for date in sorted(stats['dates'])))
            ws2.cell(row=row, column=3, value=stats['total'])
            ws2.cell(row=row, column=4, value=stats['special_days'])
            ws2.cell(row=row, column=5, value=len(stats['weeks']))

        # Adjust column widths for personal schedules
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        logging.info(f"Exported schedule to {filename}")

# Usage example
start_date = datetime(2024, 9, 28)
end_date = datetime(2024, 12, 31)
names = ["Shakir", "Fikhry", "Aiman", "Luthfi", "Dalvin", "Hazim", "Jerry", "Yassin", "Donavan"]
holidays = [datetime(2024, 10, 30), datetime(2024, 10, 31), datetime(2024, 12, 31), datetime(2024, 12, 30), datetime(2024, 12, 24), datetime(2024, 12, 25)]

# Rotate the list of names by a random number
rotation = random.randint(0, len(names) - 1)
rotated_names = rotate_list(names, rotation)
logging.info(f"Names rotated by {rotation} positions: {rotated_names}")

scheduler = ShiftScheduler(start_date, end_date, rotated_names, holidays)
scheduler.generate_schedule()
scheduler.print_schedule()
scheduler.print_personal_schedules()
scheduler.print_statistics()
scheduler.export_to_excel("shift_schedule.xlsx")
