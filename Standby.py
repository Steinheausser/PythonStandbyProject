import random
import logging
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Set up logging
log_filename = "shift_scheduler.log"
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
)

# Redirect stdout and stderr to logging
class LoggingStreamHandler:
    def __init__(self, level):
        self.level = level

    def write(self, message):
        if message.strip() != "":
            self.level(message.strip())

    def flush(self):
        pass

sys.stdout = LoggingStreamHandler(logging.info)
sys.stderr = LoggingStreamHandler(logging.error)

class ShiftScheduler:
    def __init__(self, start_date, end_date, names, holidays):
        self.start_date = start_date
        self.end_date = end_date
        self.names = names
        self.holidays = holidays
        self.schedule = {}
        self.assignments = {name: {'total': 0, 'special_days': 0, 'dates': [], 'last_assigned': None} for name in names}
        self.special_days = self.calculate_special_days()
        self.total_shifts = ((self.end_date - self.start_date).days + 1) * 2

    def calculate_special_days(self):
        special_days = []
        current_date = self.start_date
        while current_date <= self.end_date:
            if current_date.weekday() >= 5 or current_date in self.holidays:
                special_days.append(current_date)
            current_date += timedelta(days=1)
        return special_days

    def is_special_day(self, date):
        return date.weekday() >= 5 or date in self.holidays

    def generate_schedule(self):
        dates = [self.start_date + timedelta(days=i) for i in range((self.end_date - self.start_date).days + 1)]
        random.shuffle(dates)  # Randomize the order of dates

        for current_date in dates:
            available_names = self.names.copy()
            random.shuffle(available_names)

            shift = []
            for _ in range(2):  # Assign 2 people per day
                if not available_names:
                    available_names = self.names.copy()
                    random.shuffle(available_names)

                # Sort by total shifts and special days to ensure balanced distribution
                available_names.sort(key=lambda x: (self.assignments[x]['total'], self.assignments[x]['special_days']))

                name = available_names.pop(0)
                shift.append(name)
                self.assignments[name]['total'] += 1
                self.assignments[name]['dates'].append(current_date)
                self.assignments[name]['last_assigned'] = current_date

                if self.is_special_day(current_date):
                    self.assignments[name]['special_days'] += 1

            self.schedule[current_date] = shift

    def print_schedule(self):
        print("Schedule by Date:")
        for date, shift in sorted(self.schedule.items()):
            print(f"{date.strftime('%Y-%m-%d')}: {', '.join(shift)}")

    def print_personal_schedules(self):
        print("\nSchedule by Person:")
        for name, stats in self.assignments.items():
            dates = ', '.join(date.strftime('%Y-%m-%d') for date in stats['dates'])
            print(f"{name}: {dates}")

    def print_statistics(self):
        print("\nAssignment Statistics:")
        for name, stats in self.assignments.items():
            print(f"{name}: Total: {stats['total']}, Special Days: {stats['special_days']}")

        print("\nOverall Statistics:")
        print(f"Total number of standbys: {self.total_shifts}")
        print(f"Total number of Special Days: {len(self.special_days)}")

    def export_to_excel(self, filename):
        wb = Workbook()

        # Shift Schedule sheet
        ws = wb.active
        ws.title = "Shift Schedule"
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        ws['A1'] = "Date"
        ws['B1'] = "Person 1"
        ws['C1'] = "Person 2"
        ws['D1'] = "Special Day"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Data for shift schedule
        for row, (date, shift) in enumerate(sorted(self.schedule.items()), start=2):
            ws.cell(row=row, column=1, value=date.strftime("%d/%m/%Y"))
            for col, name in enumerate(shift, start=2):
                cell = ws.cell(row=row, column=col, value=name)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value="Yes" if self.is_special_day(date) else "No")
            ws.cell(row=row, column=4).border = border
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

        # Adjust column widths for shift schedule
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Personal Schedule sheet
        ws_personal = wb.create_sheet(title="Personal Schedules")
        ws_personal['A1'] = "Person"
        ws_personal['B1'] = "Dates"
        ws_personal['C1'] = "Total Shifts"
        ws_personal['D1'] = "Special Days"
        for cell in ws_personal[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        # Data for personal schedules
        for row, (name, stats) in enumerate(self.assignments.items(), start=2):
            dates = ', '.join(date.strftime('%d/%m/%Y') for date in stats['dates'])
            ws_personal.cell(row=row, column=1, value=name)
            ws_personal.cell(row=row, column=2, value=dates)
            ws_personal.cell(row=row, column=3, value=stats['total'])
            ws_personal.cell(row=row, column=4, value=stats['special_days'])
            for col in range(1, 5):
                ws_personal.cell(row=row, column=col).border = border
                ws_personal.cell(row=row, column=col).alignment = Alignment(horizontal="center")

        # Adjust column widths for personal schedules
        for column in ws_personal.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_personal.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        print(f"Schedule exported to {filename}")

# Example usage
start_date = datetime(2024, 10, 1)
end_date = datetime(2024, 12, 31)
names = ["Shakir", "Fikhry", "Aiman", "Luthfi", "Dalvin", "Hazim", "Jerry", "Yassin", "Donavan"]
holidays = [
    datetime(2024, 12, 25),  # Christmas
    datetime(2024, 12, 31),  # New Year's Eve
]
scheduler = ShiftScheduler(start_date, end_date, names, holidays)
scheduler.generate_schedule()
scheduler.print_schedule()
scheduler.print_personal_schedules()
scheduler.print_statistics()
scheduler.export_to_excel("shift_schedule.xlsx")
